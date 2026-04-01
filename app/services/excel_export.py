"""Excel export for schedule results."""
from datetime import date, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from ..models.schedule import Schedule, ScheduleEntry
from ..models.dealer import Dealer


def export_schedule_excel(schedule_id: int, db: Session) -> BytesIO:
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise ValueError("Schedule not found")

    entries = db.query(ScheduleEntry).filter(ScheduleEntry.schedule_id == schedule_id).all()
    week_start = schedule.week_start
    week_dates = [week_start + timedelta(days=i) for i in range(7)]

    # Build entry map: (dealer_id, date) -> shift
    entry_map: dict[tuple[str, date], str] = {}
    dealer_ids = set()
    for e in entries:
        entry_map[(e.dealer_id, e.date)] = e.shift
        dealer_ids.add(e.dealer_id)

    # Load dealers
    dealers = db.query(Dealer).filter(Dealer.id.in_(dealer_ids)).order_by(Dealer.last_name, Dealer.first_name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule {week_start.isoformat()}"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(week_dates))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{schedule.dealer_type.title()} Schedule: {week_start.isoformat()} - {(week_start + timedelta(days=6)).isoformat()}"
    title_cell.font = Font(bold=True, size=14)

    # Header row
    headers = ["Last Name", "First Name", "EE Number"]
    for d in week_dates:
        headers.append(f"{d.strftime('%a')}\n{d.strftime('%m/%d')}")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, dealer in enumerate(dealers, 4):
        ws.cell(row=row_idx, column=1, value=dealer.last_name).border = thin_border
        ws.cell(row=row_idx, column=2, value=dealer.first_name).border = thin_border
        ws.cell(row=row_idx, column=3, value=dealer.ee_number or dealer.id).border = thin_border

        for col_offset, day in enumerate(week_dates):
            cell = ws.cell(row=row_idx, column=4 + col_offset)
            cell.border = thin_border
            cell.alignment = center
            shift = entry_map.get((dealer.id, day))
            if shift:
                cell.value = shift
                cell.fill = green_fill
                cell.font = Font(bold=True, color="2E7D32")
            else:
                cell.value = "OFF"
                cell.fill = red_fill
                cell.font = Font(color="999999")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    for i in range(len(week_dates)):
        col_letter = chr(ord("D") + i)
        ws.column_dimensions[col_letter].width = 10

    # Summary row
    summary_row = len(dealers) + 5
    ws.cell(row=summary_row, column=1, value="Total Assigned").font = header_font
    for col_offset, day in enumerate(week_dates):
        count = sum(1 for d in dealers if (d.id, day) in entry_map)
        ws.cell(row=summary_row, column=4 + col_offset, value=count).alignment = center

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
